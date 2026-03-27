"""
Formatea los resultados de scraping en la estructura de Creacion_productos.xlsx.

Genera un Excel con:
- Sheet "Products": 18 columnas en el orden exacto de la plantilla
- Sheet "opciones": listas de validación (Formato, Idioma, Vendor)
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from services.scrap.base import MergedBook

# ── Columnas destino (orden exacto de la plantilla) ──────────────────────────

PRODUCTS_COLUMNS = [
    "Titulo",
    "Sipnosis",
    "Vendor",
    "SKU",
    "peso (kg)",
    "Precio",
    "Precio de comparacion",
    "Portada (URL)",
    "Numero de paginas",
    "Idioma",
    "Autor",
    "Editorial",
    "Formato",
    "Alto",
    "Ancho",
    "Ilustrador",
    "Categoria",
    "Subcategoria",
]

# ── Mapeo de idioma: código/variante → nombre español ────────────────────────

_IDIOMA_MAP: dict[str, str] = {
    "es": "Español", "spa": "Español", "español": "Español",
    "castellano": "Español", "spanish": "Español",
    "en": "Ingles", "eng": "Ingles", "inglés": "Ingles",
    "english": "Ingles", "ingles": "Ingles",
    "fr": "Frances", "fre": "Frances", "fra": "Frances",
    "francés": "Frances", "frances": "Frances", "french": "Frances",
    "it": "Italiano", "ita": "Italiano", "italiano": "Italiano",
    "italian": "Italiano",
    "pt": "Portugues", "por": "Portugues", "portugués": "Portugues",
    "portugues": "Portugues", "portuguese": "Portugues",
    "de": "Aleman", "ger": "Aleman", "deu": "Aleman",
    "alemán": "Aleman", "aleman": "Aleman", "german": "Aleman",
    "ru": "Ruso", "rus": "Ruso", "ruso": "Ruso", "russian": "Ruso",
    "ar": "Arabe", "ara": "Arabe", "árabe": "Arabe", "arabe": "Arabe",
    "arabic": "Arabe",
    "zh": "Chino", "chi": "Chino", "zho": "Chino", "chino": "Chino",
    "chinese": "Chino",
    "ja": "Japones", "jpn": "Japones", "japonés": "Japones",
    "japones": "Japones", "japanese": "Japones",
    "eu": "Vasco", "eus": "Vasco", "baq": "Vasco", "vasco": "Vasco",
    "basque": "Vasco",
    "gl": "Gallego", "glg": "Gallego", "gallego": "Gallego",
    "galician": "Gallego",
    "la": "Latin", "lat": "Latin", "latín": "Latin", "latin": "Latin",
    "ca": "Catalan", "cat": "Catalan", "catalán": "Catalan",
    "catalan": "Catalan",
    "ro": "Rumano", "ron": "Rumano", "rum": "Rumano", "rumano": "Rumano",
    "romanian": "Rumano",
    "nl": "Holandes", "nld": "Holandes", "dut": "Holandes",
    "holandés": "Holandes", "holandes": "Holandes", "dutch": "Holandes",
    "bg": "Bulgaro", "bul": "Bulgaro", "búlgaro": "Bulgaro",
    "bulgaro": "Bulgaro", "bulgarian": "Bulgaro",
    "el": "Griego", "gre": "Griego", "ell": "Griego", "griego": "Griego",
    "greek": "Griego",
    "pl": "Polaco", "pol": "Polaco", "polaco": "Polaco", "polish": "Polaco",
    "cs": "Checo", "ces": "Checo", "cze": "Checo", "checo": "Checo",
    "czech": "Checo",
    "sv": "Sueco", "swe": "Sueco", "sueco": "Sueco", "swedish": "Sueco",
}

# ── Mapeo de encuadernación → formato ────────────────────────────────────────

_FORMATO_MAP: dict[str, str] = {
    "hardcover": "Tapa Dura", "hard cover": "Tapa Dura",
    "tapa dura": "Tapa Dura", "cartoné": "Tapa Dura",
    "cartone": "Tapa Dura", "cartonado": "Tapa Dura",
    "hardback": "Tapa Dura",
    "paperback": "Tapa Blanda", "soft cover": "Tapa Blanda",
    "tapa blanda": "Tapa Blanda", "rústica": "Tapa Blanda",
    "rustica": "Tapa Blanda", "softcover": "Tapa Blanda",
    "book": "Tapa Blanda",
    "pocket": "Bolsillo", "bolsillo": "Bolsillo",
    "mass market paperback": "Bolsillo",
    "spiral": "Espiral", "espiral": "Espiral",
    "spiral-bound": "Espiral", "wire-o": "Espiral",
    "board book": "Tapa Dura", "board": "Tapa Dura",
    "cloth": "Tela", "tela": "Tela",
    "stapled": "Grapado", "grapado": "Grapado",
    "troquelado": "Troquelado",
    "anillas": "Anillas", "ring": "Anillas",
}

# ── Listas para la hoja "opciones" ───────────────────────────────────────────

FORMATOS = [
    "Tapa Dura", "Tapa Blanda", "Bolsillo", "Libro de lujo", "Espiral",
    "Tela", "Grapado", "Fasciculo Encuadernable", "Troquelado", "Anillas",
    "Otros",
]

IDIOMAS = [
    "Español", "Ingles", "Frances", "Italiano", "Portugues", "Aleman",
    "Bilingue (Español-Ingles)", "Bilingue (Español-Portugues)", "Vasco",
    "Gallego", "Latin", "Ruso", "Arabe", "Chino", "Japones", "Catalan",
    "Rumano", "Holandes", "Bulgaro", "Griego", "Polaco", "Checo", "Sueco",
]

VENDORS = [
    "20x2", "25 Alas", "603 La Gran Via", "7Colores", "ACLI", "Acuarell",
    "Agenda del Mar", "Alejandra Márquez Villegas", "Alejandra Mesa",
    "Alejandra Mesa González", "Alejandro Salazar Yusti", "Alexander Cossio",
    "Alexandra Castrillon", "Alianza Editorial", "Alicia Mejia",
    "Álvaro Castaño Díaz", "Álvaro González Alorda", "Amalia Londoño",
    "Amelia Amórtegui", "Andina", "Andres Felipe Arias", "Andres Romero",
    "Angosta Editores", "Aniko Villalba", "Arbitraria", "Arpegio", "Arquine",
    "Artemis Libros", "Artimaña Editorial", "As Ediciones",
    "Asociación de Editoriales Independientes de Chile", "Atarraya Editores",
    "Aulas Amigas", "Axioma Editores", "Babel", "Beatriz Ospina G", "Biblok",
    "Books for U", "Booktique", "Bukz", "Bukz Academy", "Bukz B2B",
    "Bukz Corporativo", "Bukz España", "Bukz Stationery", "Bukz USA", "Bukz.co",
    "Caballito de Acero", "Cain Press", "Calixta Editores", "Cangrejo Editores",
    "Carolina Gaviria", "Carolina Giraldo García", "Carolina Gonzáles Jiménez",
    "Carolina Jimenez", "Carolina Montoya CocoBarroco", "Carolina Pérez Botero",
    "Cartograma + NoName", "Catherine Villota", "CHACHA", "Círculo Abierto",
    "Circulo de Lectores", "Club Editores S.A.", "Códice Producciones",
    "Codiscos", "Colectivo Remitentes", "Comfama", "Cuvico", "Cypres",
    "Daniela Umaña", "Daniela Zuluaga Velez",
    "Diana López - Proyecto Medelliniando", "Diana Martínez Mona",
    "Difusora Larousse de Colombia", "Diseños Lalys", "Dos Gatos Editores",
    "DosYDos", "Duende", "Edgar David González", "Ediciones de la U",
    "Ediciones el Silencio", "Ediciones Gamma", "Ediciones Gaviota",
    "Ediciones Urano", "Ediciones Vestigio", "Editorial Anafora",
    "Editorial Eafit", "Editorial Quimbombó", "Editorial Solar",
    "Elev8 Media S.A.S", "Empoderados SAS", "Ensifera Editores",
    "Estrategia en Ventas", "Fabio Vargas Tamayo", "FCE", "FCM", "Fera",
    "Fernando Ayerbe", "Finanzas Emocionales", "Fondo Cultural Iberoamericano",
    "Frailejón Editores", "Franko Group SAS", "Fundación Casa Arcoíris",
    "Fundación Cucú", "Fundación Cultural Viento teatro", "Gavilán Sin Cola",
    "Germán Puerta", "Gloria Jaramillo, Yamid López, Julián Vásquez",
    "Grammata - Vazquez", "Grupo Editorial Planeta", "Grupo Monserrate",
    "Grupo Penta", "Happy Kiddo", "Harol Ortiz", "Harry Marin", "Health Books",
    "HiperTexto", "Holz Haus", "Hugo Jaimezurek", "Huracán Distribución",
    "Ibsen Ochoa", "Icaro Libros", "Ícono Editorial", "Idealo Pez",
    "Ignacio Arismendi Posada", "ImasD", "Independiente", "Ingenio",
    "Jaime Botero", "JARDIN PUBLICACIONES", "Juan Gonzalo Benitez Montoya",
    "Juliana Rego Editores", "Karin Muñoz Pinto", "Kate Villota", "Kocodio",
    "Kupa", "La Diligencia", "La Editora - Ana Meza",
    "La magia de empezar de 0", "La Valija de Fuego Editorial", "Laura Blair",
    "Laura F. M.", "Laura Roca", "Lavanda Editoras", "Libro Arte S.A.S",
    "Libros de Ruta", "Libros del Fuego", "Libros del Motín",
    "Lindy María Márquez Holguín", "Lobolunar", "Luisa Gómez",
    "Luisa Robledo", "Luminosa", "Luz Karime Saleme Correa", "Maeva Ediciones",
    "Mandalas para el alma", "Marcela Bouhot", "Marco Polo",
    "María Andrea Estrada", "María Antonia Sierra",
    "Maria Emma Prieto Manjarres", "Maria Pulido Alvarez", "Mattelsa",
    "McMullan Birding", "Melon", "Mesaestandar", "Milserifas", "Mo Ediciones",
    "Museo del río magdalena", "Nelly Giraldo Gil", "No vendor", "Novili",
    "Oceano", "O-Lab", "Palabra Libre", "Panamericana",
    "Paola Rueda López S.A.S.", "Penguin RandomHouse", "Pergamino Café",
    "Perla Negra", "Pilicefalo Ediciones", "Platzi S.A.S", "Plaza & Janes",
    "Poiema Publicaciones", "Postobon", "Proyectos Sin Limites",
    "Rafael Iván Botero", "Raya Editores", "Reservoir Books", "Rey Naranjo",
    "Saga Libros", "Sara Betancur", "Sebastián Betancur",
    "Secretos para contar", "Sergio Restrepo", "Siglo del Hombre",
    "Silaba Editores", "Sin Fronteras", "Sin Ocasión", "SITRA Mundo Creativo",
    "Sol Beatriz Botero", "Taller de artes de Medellín",
    "Taller de edición Rocca", "Taller Talante", "Teresita Varon",
    "Testigo Directo", "The Black Bean", "Torrealta", "Toy",
    "Tragaluz Editores", "Trigeon SAS", "UNILAT", "Union Editorial Colombia",
    "Universidad CES", "Universidad de Antioquia", "Urban",
    "Valeria Marín Pineda", "Vasquez Editores", "Verónica Abad Londoño",
    "Verso Libre", "Viiel", "Villegas Editores", "Viviana Escobar",
    "Wellness Bites", "Wilfrido Gonzalez", "word clound classies",
    "Yoni Rendón",
]


def _map_idioma(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    return _IDIOMA_MAP.get(raw.strip().lower(), raw)


def _map_formato(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    return _FORMATO_MAP.get(raw.strip().lower(), raw)


def format_creacion(books: list[MergedBook]) -> bytes:
    """Genera un Excel en formato Creacion_productos a partir de los MergedBook."""
    wb = Workbook()

    # ── Sheet "Products" ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Products"

    # Headers
    header_font = Font(bold=True)
    for col_idx, header in enumerate(PRODUCTS_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    # Data rows
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_idx, book in enumerate(books, 2):
        ws.cell(row=row_idx, column=1, value=book.titulo)
        ws.cell(row=row_idx, column=2, value=book.descripcion)
        # col 3: Vendor — vacío (se llena manualmente)
        ws.cell(row=row_idx, column=4, value=book.isbn)            # SKU
        # col 5: peso (kg) — vacío
        # col 6: Precio — vacío
        # col 7: Precio de comparacion — vacío
        ws.cell(row=row_idx, column=8, value=book.portada_url)
        ws.cell(row=row_idx, column=9, value=book.paginas)
        ws.cell(row=row_idx, column=10, value=_map_idioma(book.idioma))
        ws.cell(row=row_idx, column=11, value=book.autor)
        ws.cell(row=row_idx, column=12, value=book.editorial)
        ws.cell(row=row_idx, column=13, value=_map_formato(book.encuadernacion))
        # col 14: Alto — vacío
        # col 15: Ancho — vacío
        # col 16: Ilustrador — vacío
        ws.cell(row=row_idx, column=17, value=book.categoria)
        # col 18: Subcategoria — vacío
        # Columnas de revisión (19-21)
        ws.cell(row=row_idx, column=19, value=book.fuente_primaria)
        ws.cell(row=row_idx, column=20, value=book.campos_encontrados)
        ws.cell(row=row_idx, column=21, value=book.alertas)

        # Resaltar filas no encontradas en amarillo
        if not book.found:
            for col in range(1, 22):
                ws.cell(row=row_idx, column=col).fill = yellow_fill

    last_row = len(books) + 1

    # Headers de revisión en rojo
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(["Fuente primaria", "Campos encontrados", "Alertas"], 19):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = red_fill
        cell.font = white_bold

    # ── Sheet "opciones" ─────────────────────────────────────────────────
    ws_opt = wb.create_sheet("opciones")
    ws_opt.cell(row=1, column=1, value="Formato")
    ws_opt.cell(row=1, column=3, value="Idioma")
    ws_opt.cell(row=1, column=5, value="Vendor")
    for i, fmt in enumerate(FORMATOS, 2):
        ws_opt.cell(row=i, column=1, value=fmt)
    for i, idioma in enumerate(IDIOMAS, 2):
        ws_opt.cell(row=i, column=3, value=idioma)
    for i, vendor in enumerate(VENDORS, 2):
        ws_opt.cell(row=i, column=5, value=vendor)

    # ── Data validations (dropdowns) en Products ─────────────────────────
    if last_row >= 2:
        fmt_formula = f'"{ ",".join(FORMATOS) }"'
        dv_formato = DataValidation(type="list", formula1=fmt_formula, allow_blank=True)
        dv_formato.error = "Seleccione un formato válido"
        dv_formato.errorTitle = "Formato inválido"
        ws.add_data_validation(dv_formato)
        dv_formato.add(f"M2:M{last_row}")

        idioma_formula = f'"{ ",".join(IDIOMAS) }"'
        dv_idioma = DataValidation(type="list", formula1=idioma_formula, allow_blank=True)
        dv_idioma.error = "Seleccione un idioma válido"
        dv_idioma.errorTitle = "Idioma inválido"
        ws.add_data_validation(dv_idioma)
        dv_idioma.add(f"J2:J{last_row}")

        vendor_last = len(VENDORS) + 1
        dv_vendor = DataValidation(
            type="list",
            formula1=f"opciones!$E$2:$E${vendor_last}",
            allow_blank=True,
        )
        dv_vendor.error = "Seleccione un vendor válido"
        dv_vendor.errorTitle = "Vendor inválido"
        ws.add_data_validation(dv_vendor)
        dv_vendor.add(f"C2:C{last_row}")

    # ── Generar bytes ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
