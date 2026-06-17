"""
Router Papyser: consolida pedidos `bookstore_requests` con `status=pending`,
arma un xlsx con 2 hojas (Total + Sedes) y lo envía por WhatsApp a Leidy (Papyser).
"""
import threading
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, HTTPException
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from firebase_admin import firestore
from google.cloud.firestore_v1.base_query import FieldFilter

from services.firebase_service import get_firestore_db
from services.whatsapp_service import send_media

router = APIRouter(prefix="/api/papyser", tags=["Papyser"])

PAPYSER_WHATSAPP = "+57 4 444 6969"
PAPYSER_CONTACTO = "Leidy"

_consolidate_lock = threading.Lock()


def _agregar_items(pendings: list) -> dict:
    """De N pedidos → dict {code: {nombre, total, por_sede: {sede: cantidad}}}."""
    agregado: dict = {}
    for doc in pendings:
        data = doc.to_dict()
        sede = data.get("branch", "Sin sede")
        for item in data.get("items", []):
            code = (item.get("code") or "").strip()
            if not code:
                continue
            cant = int(item.get("quantity") or 0)
            if cant <= 0:
                continue
            slot = agregado.setdefault(code, {
                "nombre": item.get("name", ""),
                "total": 0,
                "por_sede": {},
            })
            slot["total"] += cant
            slot["por_sede"][sede] = slot["por_sede"].get(sede, 0) + cant
    return agregado


def _construir_xlsx(agregado: dict) -> bytes:
    wb = openpyxl.Workbook()

    ws_total = wb.active
    ws_total.title = "Total"
    ws_total.append(["Código", "Producto", "Cantidad"])
    for cell in ws_total[1]:
        cell.font = Font(bold=True)
    for code, info in sorted(agregado.items(), key=lambda x: x[1]["nombre"]):
        ws_total.append([code, info["nombre"], info["total"]])

    ws_sedes = wb.create_sheet("Sedes")
    sedes = sorted({s for info in agregado.values() for s in info["por_sede"]})
    ws_sedes.append(["Código", "Producto"] + sedes + ["Total"])
    for cell in ws_sedes[1]:
        cell.font = Font(bold=True)
    for code, info in sorted(agregado.items(), key=lambda x: x[1]["nombre"]):
        row = [code, info["nombre"]] + [info["por_sede"].get(s, "") for s in sedes] + [info["total"]]
        ws_sedes.append(row)

    for ws in (ws_total, ws_sedes):
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _construir_caption(agregado: dict) -> str:
    fecha = datetime.now().strftime("%d/%m/%Y")
    total_items = sum(info["total"] for info in agregado.values())
    return (
        f"Hola {PAPYSER_CONTACTO}, te paso el pedido de Bukz para hoy {fecha}.\n\n"
        f"📋 Resumen:\n"
        f"• {len(agregado)} productos distintos\n"
        f"• {total_items} unidades en total\n\n"
        f"En el archivo encuentras 2 hojas: *Total* con la cantidad consolidada por producto, "
        f"y *Sedes* con el desglose por sede.\n\n"
        f"Quedamos pendientes de tu cotización. ¡Gracias!\n\n"
        f"Equipo Bukz"
    )


@router.get("/preview")
async def preview():
    """Previsualiza qué se consolidaría sin enviar nada. Útil para confirmación del frontend."""
    db = get_firestore_db()
    pendings = list(
        db.collection("bookstore_requests")
        .where(filter=FieldFilter("status", "==", "pending"))
        .stream()
    )
    if not pendings:
        return {
            "pedidos_pendientes": 0,
            "productos_distintos": 0,
            "total_items": 0,
            "sedes": [],
        }
    agregado = _agregar_items(pendings)
    total_items = sum(info["total"] for info in agregado.values())
    sedes = sorted({(d.to_dict().get("branch") or "Sin sede") for d in pendings})
    return {
        "pedidos_pendientes": len(pendings),
        "productos_distintos": len(agregado),
        "total_items": total_items,
        "sedes": sedes,
    }


@router.post("/consolidate")
async def consolidate():
    """Consolida pendings → xlsx → WhatsApp a Papyser → marca como requested."""
    if not _consolidate_lock.acquire(blocking=False):
        raise HTTPException(status_code=409, detail="Ya hay una consolidación en curso")
    try:
        db = get_firestore_db()
        pendings = list(
            db.collection("bookstore_requests")
            .where(filter=FieldFilter("status", "==", "pending"))
            .stream()
        )
        if not pendings:
            raise HTTPException(status_code=404, detail="No hay pedidos pendientes para consolidar")

        agregado = _agregar_items(pendings)
        if not agregado:
            raise HTTPException(
                status_code=422,
                detail="Los pedidos pendientes no contienen items con código válido"
            )

        xlsx_bytes = _construir_xlsx(agregado)
        caption = _construir_caption(agregado)
        file_name = f"Pedido Bukz {datetime.now().strftime('%Y-%m-%d')}.xlsx"

        await send_media(
            numero=PAPYSER_WHATSAPP,
            file_content=xlsx_bytes,
            file_name=file_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            caption=caption,
        )

        for doc in pendings:
            doc.reference.update({
                "status": "requested",
                "consolidatedAt": firestore.SERVER_TIMESTAMP,
            })

        return {
            "ok": True,
            "pedidos_consolidados": len(pendings),
            "productos_distintos": len(agregado),
            "total_items": sum(info["total"] for info in agregado.values()),
            "archivo": file_name,
        }
    finally:
        _consolidate_lock.release()
