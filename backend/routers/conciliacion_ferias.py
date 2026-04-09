"""
Router de Conciliación de Ferias - cruza inventario enviado (Excel),
devuelto (Excel) y ventas (Shopify) para detectar pérdidas/robos en ferias.
"""
import io
import logging
from datetime import datetime

import openpyxl
import requests
from fastapi import APIRouter, HTTPException, Form, UploadFile, File
from fastapi.responses import StreamingResponse

from config import settings
from services.shopify_service import get_locations

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/conciliacion-ferias", tags=["Conciliación Ferias"])


def _graphql(query: str, variables: dict | None = None) -> dict:
    """Ejecuta query GraphQL contra Shopify y retorna data."""
    payload: dict = {"query": query}
    if variables:
        payload["variables"] = variables

    headers = settings.get_shopify_headers()
    url = settings.get_graphql_url()

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
    except Exception as e:
        logger.error(f"Error conectando a Shopify: {e}")
        raise HTTPException(status_code=503, detail=f"No se pudo conectar a Shopify: {e}")

    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="Shopify rate limit, intenta en unos segundos")

    if resp.status_code != 200:
        logger.error(f"Shopify HTTP {resp.status_code}: {resp.text[:500]}")
        raise HTTPException(status_code=502, detail=f"Shopify respondio HTTP {resp.status_code}")

    data = resp.json()

    if "errors" in data:
        error_msgs = [e.get("message", str(e)) for e in data["errors"]]
        logger.error(f"Shopify GraphQL errors: {error_msgs}")
        raise HTTPException(status_code=400, detail=f"Error Shopify: {'; '.join(error_msgs)}")

    return data.get("data", {})


# ---------------------------------------------------------------------------
# Queries
# ---------------------------------------------------------------------------

_ORDERS_QUERY = """
query($first: Int!, $after: String, $query: String) {
  orders(first: $first, after: $after, query: $query) {
    edges {
      node {
        id
        name
        createdAt
        fulfillmentOrders(first: 10) {
          edges {
            node {
              assignedLocation {
                location { id }
              }
              lineItems(first: 50) {
                edges {
                  node {
                    sku
                    totalQuantity
                  }
                }
              }
            }
          }
        }
      }
    }
    pageInfo { hasNextPage endCursor }
  }
}
"""

_PRODUCT_VARIANTS_QUERY = """
query($first: Int!, $after: String, $query: String) {
  productVariants(first: $first, after: $after, query: $query) {
    edges {
      node {
        sku
        product { title }
      }
    }
    pageInfo { hasNextPage endCursor }
  }
}
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_excel(file_bytes: bytes) -> dict[str, int]:
    """
    Lee un archivo Excel y retorna {sku: quantity}.
    Detecta columnas por nombre de header (case-insensitive):
      - SKU: sku, codigo, code
      - Cantidad: cantidad, qty, quantity, unidades
    Si no detecta headers, usa col A = SKU, col B = cantidad.
    Ignora filas con SKU vacío o cantidad no numérica.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {}

    sku_col: int | None = None
    qty_col: int | None = None

    sku_headers = {"sku", "codigo", "code", "código"}
    qty_headers = {"cantidad", "qty", "quantity", "unidades"}

    # Intentar detectar headers en la primera fila
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    if first_row:
        for cell in first_row:
            val = str(cell.value or "").strip().lower()
            if val in sku_headers:
                sku_col = cell.column
            elif val in qty_headers:
                qty_col = cell.column

    # Fallback: primeras 2 columnas
    start_row = 1
    if sku_col is not None and qty_col is not None:
        start_row = 2  # saltar header
    else:
        sku_col = 1
        qty_col = 2

    result: dict[str, int] = {}
    for row in ws.iter_rows(min_row=start_row, values_only=False):
        sku_cell = None
        qty_cell = None
        for cell in row:
            if cell.column == sku_col:
                sku_cell = cell.value
            elif cell.column == qty_col:
                qty_cell = cell.value

        sku_val = str(sku_cell).strip() if sku_cell is not None else ""
        if not sku_val:
            continue

        try:
            qty_val = int(float(str(qty_cell)))
        except (TypeError, ValueError):
            continue

        result[sku_val] = result.get(sku_val, 0) + qty_val

    wb.close()
    return result


def _get_sales_for_location(location_id: str, fecha_inicio: str, fecha_fin: str) -> dict[str, int]:
    """Obtiene ventas por SKU en una location específica."""
    sku_totals: dict[str, int] = {}
    after = None
    query_filter = f"created_at:>={fecha_inicio} created_at:<={fecha_fin}"

    while True:
        variables: dict = {"first": 50, "query": query_filter}
        if after:
            variables["after"] = after
        data = _graphql(_ORDERS_QUERY, variables)
        connection = data.get("orders", {})
        edges = connection.get("edges", [])

        for edge in edges:
            order_node = edge["node"]
            fo_edges = (order_node.get("fulfillmentOrders") or {}).get("edges", [])
            for fo_edge in fo_edges:
                fo_node = fo_edge["node"]
                assigned_loc = (fo_node.get("assignedLocation") or {}).get("location") or {}
                if assigned_loc.get("id") != location_id:
                    continue
                li_edges = (fo_node.get("lineItems") or {}).get("edges", [])
                for li_edge in li_edges:
                    li_node = li_edge["node"]
                    sku = li_node.get("sku")
                    qty = li_node.get("totalQuantity", 0)
                    if sku:
                        sku_totals[sku] = sku_totals.get(sku, 0) + qty

        page_info = connection.get("pageInfo", {})
        if not page_info.get("hasNextPage"):
            break
        after = page_info.get("endCursor")

    return sku_totals


def _enrich_titles(skus: set[str]) -> dict[str, str]:
    """Obtiene títulos de productos para un set de SKUs."""
    titles: dict[str, str] = {}
    sku_list = list(skus)

    # Batches de 20 SKUs
    batch_size = 20
    for i in range(0, len(sku_list), batch_size):
        batch = sku_list[i:i + batch_size]
        query_str = " OR ".join(f"sku:{s}" for s in batch)
        after = None
        while True:
            variables: dict = {"first": 100, "query": query_str}
            if after:
                variables["after"] = after
            data = _graphql(_PRODUCT_VARIANTS_QUERY, variables)
            connection = data.get("productVariants", {})
            for edge in connection.get("edges", []):
                node = edge["node"]
                sku = node.get("sku")
                title = (node.get("product") or {}).get("title", "")
                if sku and title:
                    titles[sku] = title
            page_info = connection.get("pageInfo", {})
            if not page_info.get("hasNextPage"):
                break
            after = page_info.get("endCursor")

    return titles


def _run_conciliacion(
    location_name: str,
    location_id: str,
    fecha_inicio: str,
    fecha_fin: str,
    enviados_por_sku: dict[str, int],
    devueltos_por_sku: dict[str, int],
    nombre_archivo_enviado: str,
    nombre_archivo_devuelto: str,
) -> dict:
    """Ejecuta toda la lógica de conciliación y retorna el resultado."""

    # Validaciones
    if not location_name or not location_id or not fecha_inicio or not fecha_fin:
        raise HTTPException(status_code=400, detail="Todos los campos son obligatorios")

    if not location_id.startswith("gid://shopify/Location/"):
        raise HTTPException(status_code=400, detail="location_id debe empezar con gid://shopify/Location/")

    try:
        dt_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        dt_fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Usar YYYY-MM-DD")

    if dt_inicio > dt_fin:
        raise HTTPException(status_code=400, detail="fecha_inicio debe ser <= fecha_fin")

    logger.info(f"Conciliación: {location_name} del {fecha_inicio} al {fecha_fin}")
    logger.info(f"Excel enviado: {len(enviados_por_sku)} SKUs, Excel devuelto: {len(devueltos_por_sku)} SKUs")

    # Ventas en la location (desde Shopify)
    vendido_por_sku = _get_sales_for_location(location_id, fecha_inicio, fecha_fin)
    logger.info(f"Ventas: {len(vendido_por_sku)} SKUs")

    # Enrichment de títulos
    all_skus = set(enviados_por_sku.keys()) | set(devueltos_por_sku.keys()) | set(vendido_por_sku.keys())
    titles = _enrich_titles(all_skus) if all_skus else {}

    # Calcular diferencias
    items = []
    total_enviado = 0
    total_devuelto = 0
    total_vendido = 0
    total_diferencia = 0
    skus_ok = 0
    skus_faltante = 0
    skus_sobrante = 0

    for sku in sorted(all_skus):
        env = enviados_por_sku.get(sku, 0)
        dev = devueltos_por_sku.get(sku, 0)
        ven = vendido_por_sku.get(sku, 0)
        diff = env - dev - ven

        if diff > 0:
            estado = "faltante"
            skus_faltante += 1
        elif diff < 0:
            estado = "sobrante"
            skus_sobrante += 1
        else:
            estado = "ok"
            skus_ok += 1

        total_enviado += env
        total_devuelto += dev
        total_vendido += ven
        total_diferencia += diff

        items.append({
            "sku": sku,
            "titulo": titles.get(sku, ""),
            "enviado": env,
            "devuelto": dev,
            "vendido": ven,
            "diferencia": diff,
            "estado": estado,
        })

    # Ordenar: primero faltantes (desc por diferencia), luego sobrantes, luego ok
    def sort_key(item):
        order = {"faltante": 0, "sobrante": 1, "ok": 2}
        return (order.get(item["estado"], 3), -abs(item["diferencia"]))

    items.sort(key=sort_key)

    return {
        "resumen": {
            "location": location_name,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_enviado": total_enviado,
            "total_devuelto": total_devuelto,
            "total_vendido": total_vendido,
            "total_diferencia": total_diferencia,
            "total_skus": len(all_skus),
            "skus_ok": skus_ok,
            "skus_faltante": skus_faltante,
            "skus_sobrante": skus_sobrante,
        },
        "items": items,
        "archivo_enviado": nombre_archivo_enviado,
        "archivo_devuelto": nombre_archivo_devuelto,
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/locations")
def list_locations():
    """Retorna las ubicaciones de Shopify."""
    try:
        locations_dict = get_locations()
        locations = [{"name": name, "id": gid} for name, gid in locations_dict.items()]
        return {"locations": locations}
    except Exception as e:
        logger.error(f"Error obteniendo locations: {e}")
        raise HTTPException(status_code=502, detail=f"Error obteniendo locations de Shopify: {e}")


@router.post("/conciliar")
async def conciliar(
    file_enviado: UploadFile = File(...),
    file_devuelto: UploadFile = File(...),
    location_name: str = Form(...),
    location_id: str = Form(...),
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(...),
):
    """Ejecuta la conciliación y retorna los resultados."""
    enviado_bytes = await file_enviado.read()
    devuelto_bytes = await file_devuelto.read()

    enviados_por_sku = _parse_excel(enviado_bytes)
    devueltos_por_sku = _parse_excel(devuelto_bytes)

    if not enviados_por_sku:
        raise HTTPException(status_code=400, detail="El archivo de inventario enviado está vacío o no se pudo leer")

    if not devueltos_por_sku:
        raise HTTPException(status_code=400, detail="El archivo de inventario devuelto está vacío o no se pudo leer")

    return _run_conciliacion(
        location_name=location_name,
        location_id=location_id,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        enviados_por_sku=enviados_por_sku,
        devueltos_por_sku=devueltos_por_sku,
        nombre_archivo_enviado=file_enviado.filename or "enviado.xlsx",
        nombre_archivo_devuelto=file_devuelto.filename or "devuelto.xlsx",
    )


@router.post("/exportar")
async def exportar(
    file_enviado: UploadFile = File(...),
    file_devuelto: UploadFile = File(...),
    location_name: str = Form(...),
    location_id: str = Form(...),
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(...),
):
    """Ejecuta la conciliación y genera un Excel con los resultados."""
    enviado_bytes = await file_enviado.read()
    devuelto_bytes = await file_devuelto.read()

    enviados_por_sku = _parse_excel(enviado_bytes)
    devueltos_por_sku = _parse_excel(devuelto_bytes)

    if not enviados_por_sku:
        raise HTTPException(status_code=400, detail="El archivo de inventario enviado está vacío o no se pudo leer")

    if not devueltos_por_sku:
        raise HTTPException(status_code=400, detail="El archivo de inventario devuelto está vacío o no se pudo leer")

    result = _run_conciliacion(
        location_name=location_name,
        location_id=location_id,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        enviados_por_sku=enviados_por_sku,
        devueltos_por_sku=devueltos_por_sku,
        nombre_archivo_enviado=file_enviado.filename or "enviado.xlsx",
        nombre_archivo_devuelto=file_devuelto.filename or "devuelto.xlsx",
    )

    resumen = result["resumen"]
    items = result["items"]
    faltantes = [i for i in items if i["estado"] == "faltante"]

    wb = openpyxl.Workbook()

    # Hoja Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    resumen_rows = [
        ("Campo", "Valor"),
        ("Feria", resumen["location"]),
        ("Fecha Inicio", resumen["fecha_inicio"]),
        ("Fecha Fin", resumen["fecha_fin"]),
        ("Total Enviado", resumen["total_enviado"]),
        ("Total Devuelto", resumen["total_devuelto"]),
        ("Total Vendido", resumen["total_vendido"]),
        ("Diferencia Total", resumen["total_diferencia"]),
        ("Total SKUs", resumen["total_skus"]),
        ("SKUs OK", resumen["skus_ok"]),
        ("SKUs Faltantes", resumen["skus_faltante"]),
        ("SKUs Sobrantes", resumen["skus_sobrante"]),
        ("Archivo Enviado", result["archivo_enviado"]),
        ("Archivo Devuelto", result["archivo_devuelto"]),
    ]
    for row in resumen_rows:
        ws_resumen.append(row)
    # Ajustar ancho
    ws_resumen.column_dimensions["A"].width = 20
    ws_resumen.column_dimensions["B"].width = 30

    # Hoja Detalle
    ws_detalle = wb.create_sheet("Detalle")
    headers = ["SKU", "Título", "Enviado", "Devuelto", "Vendido", "Diferencia", "Estado"]
    ws_detalle.append(headers)
    for item in items:
        ws_detalle.append([
            item["sku"],
            item["titulo"],
            item["enviado"],
            item["devuelto"],
            item["vendido"],
            item["diferencia"],
            item["estado"],
        ])
    # Ajustar anchos
    ws_detalle.column_dimensions["A"].width = 18
    ws_detalle.column_dimensions["B"].width = 40
    for col in ["C", "D", "E", "F"]:
        ws_detalle.column_dimensions[col].width = 12
    ws_detalle.column_dimensions["G"].width = 12

    # Hoja Faltantes
    ws_faltantes = wb.create_sheet("Faltantes")
    ws_faltantes.append(headers)
    for item in faltantes:
        ws_faltantes.append([
            item["sku"],
            item["titulo"],
            item["enviado"],
            item["devuelto"],
            item["vendido"],
            item["diferencia"],
            item["estado"],
        ])
    ws_faltantes.column_dimensions["A"].width = 18
    ws_faltantes.column_dimensions["B"].width = 40
    for col in ["C", "D", "E", "F"]:
        ws_faltantes.column_dimensions[col].width = 12
    ws_faltantes.column_dimensions["G"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    location_clean = resumen["location"].replace(" ", "_")
    filename = f"Conciliacion_{location_clean}_{resumen['fecha_inicio']}_{resumen['fecha_fin']}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
