"""
Router: Stock Muerto por Proveedor
Endpoints:
  POST /api/dead-stock/start   -> Inicia analisis en background
  GET  /api/dead-stock/status  -> Estado y resultados

Usa Bulk Operations de Shopify para evitar throttling de rate-limit.
Las Bulk Ops se ejecutan asincrónicamente en servidores de Shopify y no
compiten por el bucket de puntos con otros queries.
"""

import base64
import json
import threading
import time
import requests
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import APIRouter
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import settings
from services.shopify_service import check_bulk_operation_status

router = APIRouter(prefix="/api/dead-stock", tags=["Dead Stock"])

# -- Estado global del job -----------------------------------------------

_job_lock = threading.Lock()
_job: dict = {
    "running": False,
    "phase": None,
    "error": None,
    "result": None,
    "started_at": None,
}

TARGET_LOCATIONS = [
    "Bukz Las Lomas",
    "Bukz Bogota 109",
    "Bukz Viva Envigado",
    "Bukz Museo de Antioquia",
]


# -- Request model -------------------------------------------------------

class DeadStockRequest(BaseModel):
    vendor: str
    days_without_sales: int = 90
    min_product_age_months: int = 2


# -- Bulk Operation helpers -----------------------------------------------

def _wait_for_free_bulk_slot(max_wait: int = 120):
    """Espera a que no haya ninguna Bulk Operation activa en Shopify."""
    deadline = time.monotonic() + max_wait
    while time.monotonic() < deadline:
        status = check_bulk_operation_status()
        if status.get("status") not in ("RUNNING", "CREATED"):
            return
        print("[DEAD-STOCK] Esperando que termine Bulk Op en curso...", flush=True)
        time.sleep(5)
    raise RuntimeError("Timeout esperando a que termine la Bulk Operation activa")


def _start_bulk_op(inner_query: str) -> str:
    """
    Lanza una Bulk Operation con la query dada.
    Espera si hay otra bulk op activa. Retorna operation_id.
    """
    _wait_for_free_bulk_slot()

    mutation = """
    mutation {
      bulkOperationRunQuery(
        query: \"\"\"%s\"\"\"
      ) {
        bulkOperation { id status }
        userErrors { field message }
      }
    }
    """ % inner_query

    resp = requests.post(
        settings.get_graphql_url(),
        json={"query": mutation},
        headers=settings.get_shopify_headers(),
        timeout=30,
    )
    resp.raise_for_status()
    body = resp.json()

    if "errors" in body:
        raise RuntimeError(f"GraphQL errors al iniciar Bulk Op: {body['errors']}")

    result = body.get("data", {}).get("bulkOperationRunQuery", {})
    user_errors = result.get("userErrors", [])
    if user_errors:
        msgs = "; ".join(e.get("message", "") for e in user_errors)
        raise RuntimeError(f"Bulk Op userErrors: {msgs}")

    op = result.get("bulkOperation", {})
    op_id = op.get("id")
    if not op_id:
        raise RuntimeError("Bulk Op no retorno operation ID")

    print(f"[DEAD-STOCK] Bulk Op iniciada: {op_id} ({op.get('status')})", flush=True)
    return op_id


def _poll_bulk_op(op_id: str, max_wait: int = 600) -> str:
    """
    Poll una bulk operation especifica por ID hasta COMPLETED.
    Usar node(id:) en vez de currentBulkOperation evita confundir
    nuestra op con otra del scheduler o reposiciones.
    """
    # Poll por ID especifico (mas seguro que currentBulkOperation)
    query = """
    {
      node(id: "%s") {
        ... on BulkOperation {
          id status errorCode objectCount url
        }
      }
    }
    """ % op_id
    headers = settings.get_shopify_headers()
    graphql_url = settings.get_graphql_url()
    start = time.monotonic()

    while time.monotonic() - start < max_wait:
        time.sleep(5)
        try:
            resp = requests.post(
                graphql_url, json={"query": query}, headers=headers, timeout=30,
            )
            if resp.status_code != 200:
                continue
            op = resp.json().get("data", {}).get("node")
            if not op:
                continue

            status = op.get("status")
            count = op.get("objectCount", 0)
            elapsed = int(time.monotonic() - start)

            if elapsed % 30 < 6:
                print(
                    f"[DEAD-STOCK] Bulk Op {op_id}: status={status}, "
                    f"objects={count}, elapsed={elapsed}s",
                    flush=True,
                )

            if status == "COMPLETED":
                download_url = op.get("url")
                print(
                    f"[DEAD-STOCK] Bulk Op completada: {count} objetos, "
                    f"url={'SI' if download_url else 'NULL (0 resultados)'}",
                    flush=True,
                )
                return download_url  # Puede ser None si 0 resultados

            if status in ("FAILED", "CANCELED"):
                raise RuntimeError(
                    f"Bulk Op {status}: {op.get('errorCode', 'unknown')}"
                )
        except RuntimeError:
            raise
        except Exception as e:
            print(f"[DEAD-STOCK] Error polling Bulk Op: {e}", flush=True)

    raise RuntimeError(f"Bulk Op timeout despues de {max_wait}s")


def _download_jsonl(url: str) -> list[dict]:
    """Descarga JSONL de Shopify y retorna lista de objetos parseados."""
    resp = requests.get(url, timeout=300, stream=True)
    resp.raise_for_status()
    objects = []
    for line in resp.iter_lines():
        if not line:
            continue
        try:
            obj = json.loads(line.decode("utf-8") if isinstance(line, bytes) else line)
            objects.append(obj)
        except (json.JSONDecodeError, UnicodeDecodeError):
            continue
    return objects


# -- Fase 1: Obtener productos del vendor (Bulk Operation) ----------------

def _fetch_vendor_products(vendor: str, min_age_months: int) -> list[dict]:
    """
    Usa Bulk Operation para obtener todos los productos del vendor con
    inventario por location. Filtra por antiguedad y stock > 0.
    Retorna lista de variantes con el mismo formato que la version original.
    """
    cutoff_date = datetime.now() - timedelta(days=min_age_months * 30)
    safe_vendor = vendor.replace("'", "\\'").replace('"', '\\"')

    inner_query = """
        {
          products(query: "vendor:'%s'") {
            edges {
              node {
                id
                title
                vendor
                createdAt
                variants {
                  edges {
                    node {
                      id
                      sku
                      barcode
                      title
                      inventoryItem {
                        id
                        inventoryLevels {
                          edges {
                            node {
                              id
                              location {
                                name
                              }
                              quantities(names: ["available"]) {
                                name
                                quantity
                              }
                            }
                          }
                        }
                      }
                    }
                  }
                }
              }
            }
          }
        }
    """ % safe_vendor

    print(f"[DEAD-STOCK] Iniciando Bulk Op de productos para '{vendor}'...", flush=True)
    op_id = _start_bulk_op(inner_query)
    download_url = _poll_bulk_op(op_id, max_wait=600)

    # URL null = 0 resultados de la bulk op
    if not download_url:
        print("[DEAD-STOCK] Bulk Op completada con 0 resultados (URL null)", flush=True)
        return []

    objects = _download_jsonl(download_url)
    print(f"[DEAD-STOCK] Descargados {len(objects)} registros JSONL", flush=True)

    # -- Parsear JSONL plano con __parentId --
    # Bulk Ops aplastan la jerarquia: cada connection node es una linea separada.
    # Single-object relations (como inventoryItem) se inlinean en el parent.
    #
    # Estructura esperada:
    #   Product:          {id: "gid://.../Product/X", title, vendor, createdAt}
    #   ProductVariant:   {id: "gid://.../ProductVariant/Y", sku, barcode, title,
    #                      inventoryItem: {id: "gid://.../InventoryItem/Z"},
    #                      __parentId: "gid://.../Product/X"}
    #   InventoryLevel:   {id: "gid://.../InventoryLevel/W", location: {name},
    #                      quantities: [...],
    #                      __parentId: variant_gid o inventoryItem_gid}

    products_by_id: dict[str, dict] = {}
    variants_by_id: dict[str, dict] = {}
    inv_item_to_variant: dict[str, str] = {}  # InventoryItem gid -> Variant gid
    variants_result: list[dict] = []

    # Primera pasada: construir mapeos de Product, Variant, InventoryItem
    for obj in objects:
        gid = obj.get("id", "")
        parent_id = obj.get("__parentId", "")

        if "Product/" in gid and "ProductVariant/" not in gid:
            products_by_id[gid] = {
                "title": (obj.get("title") or "").strip(),
                "vendor": (obj.get("vendor") or "").strip(),
                "createdAt": obj.get("createdAt", ""),
            }

        elif "ProductVariant/" in gid:
            variants_by_id[gid] = {
                "sku": (obj.get("sku") or "").strip(),
                "barcode": (obj.get("barcode") or "").strip(),
                "title": (obj.get("title") or "").strip(),
                "product_gid": parent_id,
            }
            inv_item = obj.get("inventoryItem") or {}
            if inv_item.get("id"):
                inv_item_to_variant[inv_item["id"]] = gid

        elif "InventoryItem/" in gid:
            # InventoryItem como linea separada (no inline)
            if parent_id and "ProductVariant/" in parent_id:
                inv_item_to_variant[gid] = parent_id

    # Segunda pasada: procesar InventoryLevels
    for obj in objects:
        gid = obj.get("id", "")
        parent_id = obj.get("__parentId", "")

        if "InventoryLevel/" not in gid:
            continue

        loc_name = (obj.get("location") or {}).get("name", "")
        if loc_name not in TARGET_LOCATIONS:
            continue

        qty = 0
        for q in obj.get("quantities", []):
            if q.get("name") == "available":
                qty = q.get("quantity", 0)
                break
        if qty <= 0:
            continue

        # Resolver el variant parent (puede ser directo o via InventoryItem)
        variant_gid = None
        if parent_id in variants_by_id:
            variant_gid = parent_id
        elif parent_id in inv_item_to_variant:
            variant_gid = inv_item_to_variant[parent_id]

        if not variant_gid or variant_gid not in variants_by_id:
            continue

        variant = variants_by_id[variant_gid]
        product = products_by_id.get(variant["product_gid"], {})

        # Filtrar por antiguedad del producto
        created_str = product.get("createdAt", "")
        if created_str:
            try:
                created_at = datetime.fromisoformat(
                    created_str.replace("Z", "+00:00")
                )
                if created_at.replace(tzinfo=None) > cutoff_date:
                    continue
            except (ValueError, TypeError):
                pass

        variants_result.append({
            "sku": variant["sku"],
            "barcode": variant["barcode"],
            "product_title": product.get("title", ""),
            "variant_title": variant["title"],
            "vendor": product.get("vendor", vendor),
            "created_at": created_str[:10] if created_str else "",
            "location": loc_name,
            "stock": qty,
        })

    # Diagnostico detallado
    inv_levels_total = sum(1 for o in objects if "InventoryLevel/" in o.get("id", ""))
    inv_levels_at_target = sum(
        1 for o in objects
        if "InventoryLevel/" in o.get("id", "")
        and (o.get("location") or {}).get("name", "") in TARGET_LOCATIONS
    )
    print(
        f"[DEAD-STOCK] Parseo: {len(products_by_id)} productos, "
        f"{len(variants_by_id)} variantes, "
        f"{len(inv_item_to_variant)} inv_items mapeados, "
        f"{inv_levels_total} inv_levels total, "
        f"{inv_levels_at_target} en sedes objetivo, "
        f"{len(variants_result)} variantes finales con stock",
        flush=True,
    )

    return variants_result


# -- Fase 2: Obtener SKUs vendidos en el periodo (Bulk Operation) ---------

def _fetch_sold_skus(days: int) -> set[str]:
    """
    Usa Bulk Operation para obtener todas las ordenes pagadas en el periodo.
    Retorna set de SKUs que tuvieron al menos 1 venta.
    """
    date_start = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%dT00:00:00Z")

    inner_query = """
        {
          orders(query: "created_at:>=%s financial_status:paid") {
            edges {
              node {
                id
                lineItems {
                  edges {
                    node {
                      sku
                    }
                  }
                }
              }
            }
          }
        }
    """ % date_start

    print(f"[DEAD-STOCK] Iniciando Bulk Op de ventas (ultimos {days} dias)...", flush=True)
    op_id = _start_bulk_op(inner_query)
    download_url = _poll_bulk_op(op_id, max_wait=600)

    if not download_url:
        print("[DEAD-STOCK] Bulk Op ventas: 0 resultados (URL null)", flush=True)
        return set()

    objects = _download_jsonl(download_url)
    print(f"[DEAD-STOCK] Descargados {len(objects)} registros JSONL de ventas", flush=True)

    # Parsear: solo necesitamos SKUs de LineItems
    sold_skus: set[str] = set()
    for obj in objects:
        sku = (obj.get("sku") or "").strip()
        if sku:
            sold_skus.add(sku)

    print(
        f"[DEAD-STOCK] Ventas: {len(sold_skus)} SKUs vendidos en ultimos {days} dias",
        flush=True,
    )
    return sold_skus


# -- Fase 3: Cruce y generacion de Excel ---------------------------------

def _generate_excel(dead_rows: list[dict], vendor: str, days: int) -> str:
    """Genera Excel con los productos sin ventas y retorna base64."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Muerto"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "SKU",
        "Codigo de Barras",
        "Producto",
        "Variante",
        "Proveedor",
        "Fecha Creacion",
        "Sede",
        "Stock Disponible",
        f"Dias Sin Venta (>{days}d)",
        "Estado",
    ]

    # Escribir headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir datos
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    for i, row in enumerate(dead_rows, 2):
        values = [
            row["sku"],
            row["barcode"],
            row["product_title"],
            row["variant_title"],
            row["vendor"],
            row["created_at"],
            row["location"],
            row["stock"],
            f">{days}",
            "Stock Muerto",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col == 10:  # Columna Estado
                cell.fill = red_fill

    # Ajustar anchos de columna
    col_widths = [15, 18, 45, 20, 20, 14, 22, 16, 18, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Autofilter
    ws.auto_filter.ref = f"A1:J{len(dead_rows) + 1}"

    # Congelar header
    ws.freeze_panes = "A2"

    # Guardar a buffer y encodear base64
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


# -- Background worker ---------------------------------------------------

def _dead_stock_worker(vendor: str, days: int, min_age_months: int):
    """Ejecuta el analisis de stock muerto en 3 fases usando Bulk Operations."""
    try:
        # Fase 1: Productos del vendor con stock (Bulk Operation)
        _job["phase"] = "products"
        print(f"[DEAD-STOCK] === FASE 1: Productos de '{vendor}' ===", flush=True)
        vendor_variants = _fetch_vendor_products(vendor, min_age_months)

        if not vendor_variants:
            _job["result"] = {
                "vendor": vendor,
                "days_without_sales": days,
                "min_product_age_months": min_age_months,
                "total_variants_with_stock": 0,
                "dead_stock_variants": 0,
                "dead_stock_units": 0,
                "dead_stock_pct": 0,
                "excel_base64": None,
                "fecha_calculo": datetime.now().isoformat(),
                "message": f"No se encontraron productos del proveedor '{vendor}' con stock (creados hace mas de {min_age_months} meses)",
            }
            _job["error"] = None
            print(f"[DEAD-STOCK] Sin productos con stock para '{vendor}'", flush=True)
            return

        # Fase 2: SKUs vendidos (Bulk Operation)
        _job["phase"] = "sales"
        print(f"[DEAD-STOCK] === FASE 2: Ventas ultimos {days} dias ===", flush=True)
        sold_skus = _fetch_sold_skus(days)

        # Fase 3: Cruce y Excel
        _job["phase"] = "processing"
        print("[DEAD-STOCK] === FASE 3: Cruzando datos y generando Excel ===", flush=True)

        # Filtrar: variantes con stock que NO tienen ventas
        dead_rows = []
        for v in vendor_variants:
            sku = v["sku"]
            # Si no tiene SKU, no podemos verificar ventas — incluir como dead stock
            if not sku or sku not in sold_skus:
                dead_rows.append(v)

        # Ordenar por sede, luego por stock descendente
        dead_rows.sort(key=lambda r: (r["location"], -r["stock"]))

        total_variants = len(vendor_variants)
        dead_count = len(dead_rows)
        dead_units = sum(r["stock"] for r in dead_rows)
        total_units = sum(v["stock"] for v in vendor_variants)
        dead_pct = round((dead_count / total_variants) * 100, 1) if total_variants > 0 else 0

        # Generar Excel
        excel_b64 = _generate_excel(dead_rows, vendor, days) if dead_rows else None

        _job["result"] = {
            "vendor": vendor,
            "days_without_sales": days,
            "min_product_age_months": min_age_months,
            "total_variants_with_stock": total_variants,
            "total_units": total_units,
            "dead_stock_variants": dead_count,
            "dead_stock_units": dead_units,
            "dead_stock_pct": dead_pct,
            "excel_base64": excel_b64,
            "fecha_calculo": datetime.now().isoformat(),
        }
        _job["error"] = None

        print(
            f"[DEAD-STOCK] Completado: {dead_count}/{total_variants} variantes sin ventas "
            f"({dead_units} unidades, {dead_pct}%)",
            flush=True,
        )

    except Exception as e:
        _job["error"] = str(e)
        print(f"[DEAD-STOCK] Error: {e}", flush=True)
    finally:
        _job["running"] = False
        _job["phase"] = None


# -- Endpoints -----------------------------------------------------------

@router.post("/start")
def start_dead_stock(req: DeadStockRequest):
    """
    Inicia analisis de stock muerto para un proveedor en background.
    Consultar GET /status para ver progreso y resultados.
    """
    with _job_lock:
        if _job["running"]:
            return {
                "success": False,
                "error": "Ya hay un analisis en progreso",
                "phase": _job["phase"],
            }

        _job["running"] = True
        _job["error"] = None
        _job["result"] = None
        _job["started_at"] = datetime.now().isoformat()

    thread = threading.Thread(
        target=_dead_stock_worker,
        args=(req.vendor, req.days_without_sales, req.min_product_age_months),
        daemon=True,
    )
    thread.start()

    return {
        "success": True,
        "message": f"Analisis iniciado para '{req.vendor}' ({req.days_without_sales} dias, productos >{req.min_product_age_months} meses)",
    }


@router.get("/status")
def dead_stock_status():
    """Retorna el estado del analisis y los resultados si estan listos."""
    return {
        "running": _job["running"],
        "phase": _job["phase"],
        "error": _job["error"],
        "started_at": _job["started_at"],
        "result": _job["result"],
    }
