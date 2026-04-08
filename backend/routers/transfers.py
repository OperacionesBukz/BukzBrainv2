"""
Router de Transfers - consulta transfers de inventario entre ubicaciones en Shopify.
Usa GraphQL Admin API (inventoryTransfers query).
Endpoints sin auth Firebase para acceso directo (lectura solamente).
"""
import io
import logging

import openpyxl
import requests
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/transfers", tags=["Transfers"])


def _graphql(query: str, variables: dict | None = None) -> dict:
    """Ejecuta query GraphQL contra Shopify y retorna data."""
    payload: dict = {"query": query}
    if variables:
        payload["variables"] = variables

    headers = settings.get_shopify_headers()
    url = settings.get_graphql_url()

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
    except Exception as e:
        logger.error(f"Error conectando a Shopify: {e}")
        raise HTTPException(status_code=503, detail=f"No se pudo conectar a Shopify: {e}")

    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="Shopify rate limit, intenta en unos segundos")

    if resp.status_code != 200:
        logger.error(f"Shopify HTTP {resp.status_code}: {resp.text[:500]}")
        raise HTTPException(status_code=502, detail=f"Shopify respondio HTTP {resp.status_code}")

    data = resp.json()

    if "errors" in data:
        error_msgs = [e.get("message", str(e)) for e in data["errors"]]
        logger.error(f"Shopify GraphQL errors: {error_msgs}")
        raise HTTPException(status_code=400, detail=f"Error Shopify: {'; '.join(error_msgs)}")

    return data.get("data", {})


# ---------------------------------------------------------------------------
# Queries
# ---------------------------------------------------------------------------

_TRANSFERS_LIST = """
query inventoryTransfers($first: Int!, $after: String, $query: String, $sortKey: TransferSortKeys, $reverse: Boolean) {
  inventoryTransfers(first: $first, after: $after, query: $query, sortKey: $sortKey, reverse: $reverse) {
    edges {
      node {
        id
        name
        status
        dateCreated
        totalQuantity
        receivedQuantity
        note
        referenceName
        tags
        origin {
          location { id name }
        }
        destination {
          location { id name }
        }
        lineItemsCount { count }
      }
      cursor
    }
    pageInfo {
      hasNextPage
      endCursor
    }
  }
}
"""

_TRANSFER_DETAIL = """
query inventoryTransfer($id: ID!) {
  inventoryTransfer(id: $id) {
    id
    name
    status
    dateCreated
    totalQuantity
    receivedQuantity
    note
    referenceName
    tags
    origin {
      location { id name }
    }
    destination {
      location { id name }
    }
    lineItems(first: 250) {
      edges {
        node {
          id
          totalQuantity
          shippedQuantity
          inventoryItem {
            id
            sku
          }
        }
      }
      pageInfo {
        hasNextPage
        endCursor
      }
    }
  }
}
"""


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/")
def list_transfers(
    limit: int = Query(50, ge=1, le=250),
    after: str | None = Query(None, description="Cursor para paginacion"),
    query: str | None = Query(None, description="Filtro: status, origin_id, destination_id, tag, created_at"),
    sort_key: str = Query("CREATED_AT", description="Campo de ordenamiento: ID, CREATED_AT"),
    reverse: bool = Query(True, description="Orden descendente (mas recientes primero)"),
):
    """Lista transfers de inventario entre ubicaciones."""
    variables: dict = {
        "first": limit,
        "sortKey": sort_key,
        "reverse": reverse,
    }
    if after:
        variables["after"] = after
    if query:
        variables["query"] = query

    data = _graphql(_TRANSFERS_LIST, variables)
    connection = data.get("inventoryTransfers", {})
    edges = connection.get("edges", [])
    page_info = connection.get("pageInfo", {})

    transfers = []
    for edge in edges:
        node = edge["node"]
        transfers.append({
            "id": node.get("id"),
            "name": node.get("name"),
            "status": node.get("status"),
            "date_created": node.get("dateCreated"),
            "total_quantity": node.get("totalQuantity"),
            "received_quantity": node.get("receivedQuantity"),
            "note": node.get("note"),
            "reference_name": node.get("referenceName"),
            "tags": node.get("tags", []),
            "origin": (node.get("origin") or {}).get("location", {}).get("name"),
            "destination": (node.get("destination") or {}).get("location", {}).get("name"),
            "line_items_count": (node.get("lineItemsCount") or {}).get("count", 0),
            "cursor": edge.get("cursor"),
        })

    return {
        "transfers": transfers,
        "total": len(transfers),
        "has_next_page": page_info.get("hasNextPage", False),
        "end_cursor": page_info.get("endCursor"),
    }


@router.get("/debug/{transfer_id}")
def debug_transfer(transfer_id: str, sku_filter: str | None = Query(None)):
    """Debug: consulta shipment DRAFT por GID y devuelve respuesta cruda."""
    if not transfer_id.startswith("gid://"):
        transfer_id = f"gid://shopify/InventoryTransfer/{transfer_id}"

    # Paso 1: obtener shipment GIDs
    q1 = """
    query ($id: ID!) {
      inventoryTransfer(id: $id) {
        shipments(first: 10) {
          edges { node { id name status lineItemsCount { count } } }
        }
      }
    }
    """
    payload = {"query": q1, "variables": {"id": transfer_id}}
    headers = settings.get_shopify_headers()
    url = settings.get_graphql_url()
    r1 = requests.post(url, json=payload, headers=headers, timeout=30).json()

    shipments_info = []
    draft_gid = None
    for edge in (r1.get("data", {}).get("inventoryTransfer") or {}).get("shipments", {}).get("edges", []):
        s = edge["node"]
        shipments_info.append(s)
        if s.get("status") == "DRAFT":
            draft_gid = s["id"]

    if not draft_gid:
        return {"error": "No DRAFT shipment found", "shipments": shipments_info}

    # Paso 2: query directo al shipment por GID
    q2 = """
    query ($id: ID!, $first: Int!) {
      node(id: $id) {
        ... on InventoryShipment {
          id
          name
          status
          lineItems(first: $first) {
            edges {
              node {
                id
                quantity
                inventoryItem { id sku }
              }
            }
            pageInfo { hasNextPage endCursor }
          }
        }
      }
    }
    """
    r2 = requests.post(url, json={"query": q2, "variables": {"id": draft_gid, "first": 250}}, headers=headers, timeout=30).json()

    # Si hay filtro de SKU, buscar solo ese
    if sku_filter:
        matched = []
        for edge in (r2.get("data", {}).get("node") or {}).get("lineItems", {}).get("edges", []):
            n = edge["node"]
            if (n.get("inventoryItem") or {}).get("sku") == sku_filter:
                matched.append(n)
        return {"shipments": shipments_info, "draft_gid": draft_gid, "sku_filter": sku_filter, "matches": matched}

    return {"shipments": shipments_info, "draft_gid": draft_gid, "raw_response": r2}


_SHIPMENT_LINE_ITEMS_PAGE = """
query shipmentLineItems($id: ID!, $shipmentIdx: Int!, $first: Int!, $after: String) {
  inventoryTransfer(id: $id) {
    shipments(first: $shipmentIdx) {
      edges {
        node {
          name
          status
          lineItems(first: $first, after: $after) {
            edges {
              node {
                quantity
                inventoryItem { sku }
              }
            }
            pageInfo { hasNextPage endCursor }
          }
        }
      }
    }
  }
}
"""


@router.get("/export/{transfer_id}")
def export_shipment_excel(
    transfer_id: str,
    shipment_status: str = Query("DRAFT", description="Status del shipment a exportar: DRAFT, RECEIVED, etc."),
):
    """Genera Excel (SKU + QUANTITY) de un shipment especifico dentro de un transfer."""
    if not transfer_id.startswith("gid://"):
        transfer_id = f"gid://shopify/InventoryTransfer/{transfer_id}"

    # Paso 1: obtener solo IDs y status de shipments (SIN line items)
    shipments_query = """
    query inventoryTransfer($id: ID!) {
      inventoryTransfer(id: $id) {
        name
        shipments(first: 10) {
          edges {
            node { id name status }
          }
        }
      }
    }
    """
    data = _graphql(shipments_query, {"id": transfer_id})
    transfer = data.get("inventoryTransfer")
    if not transfer:
        raise HTTPException(status_code=404, detail="Transfer no encontrado")

    shipment_gid = None
    shipment_name = "shipment"
    for edge in transfer.get("shipments", {}).get("edges", []):
        s = edge["node"]
        if s.get("status") == shipment_status.upper():
            shipment_gid = s["id"]
            shipment_name = s.get("name", shipment_name)
            break

    if not shipment_gid:
        raise HTTPException(status_code=404, detail=f"No se encontro shipment con status {shipment_status}")

    # Paso 2: consultar SOLO ese shipment por su GID, paginando line items
    items: list[tuple[str, int]] = []
    after = None
    while True:
        page_query = """
        query shipmentPage($id: ID!, $first: Int!, $after: String) {
          node(id: $id) {
            ... on InventoryShipment {
              lineItems(first: $first, after: $after) {
                edges {
                  node {
                    quantity
                    inventoryItem { sku }
                  }
                }
                pageInfo { hasNextPage endCursor }
              }
            }
          }
        }
        """
        variables: dict = {"id": shipment_gid, "first": 250}
        if after:
            variables["after"] = after

        page_data = _graphql(page_query, variables)
        li_data = (page_data.get("node") or {}).get("lineItems", {})
        for edge in li_data.get("edges", []):
            n = edge["node"]
            sku = (n.get("inventoryItem") or {}).get("sku")
            qty = n.get("quantity", 0)
            if sku:
                items.append((sku, qty))

        pi = li_data.get("pageInfo", {})
        if not pi.get("hasNextPage"):
            break
        after = pi.get("endCursor")

    # Paso 3: generar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws["A1"] = "SKU"
    ws["B1"] = "QUANTITY"
    for i, (sku, qty) in enumerate(items, start=2):
        ws[f"A{i}"] = sku
        ws[f"B{i}"] = qty

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    clean_name = shipment_name.replace("#", "")
    filename = f"Transfer_{clean_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_TRANSFER_LINE_ITEMS_PAGE = """
query transferLineItems($id: ID!, $first: Int!, $after: String) {
  inventoryTransfer(id: $id) {
    lineItems(first: $first, after: $after) {
      edges {
        node {
          id
          totalQuantity
          shippedQuantity
          inventoryItem {
            id
            sku
          }
        }
      }
      pageInfo {
        hasNextPage
        endCursor
      }
    }
  }
}
"""


@router.get("/{transfer_id}")
def get_transfer(transfer_id: str):
    """Obtiene un transfer con TODOS sus line items (pagina internamente)."""
    if not transfer_id.startswith("gid://"):
        transfer_id = f"gid://shopify/InventoryTransfer/{transfer_id}"

    data = _graphql(_TRANSFER_DETAIL, {"id": transfer_id})
    transfer = data.get("inventoryTransfer")
    if not transfer:
        raise HTTPException(status_code=404, detail=f"Transfer {transfer_id} no encontrado")

    # Recopilar line items de la primera pagina
    line_items = []
    li_data = transfer.get("lineItems", {})
    for edge in li_data.get("edges", []):
        node = edge["node"]
        inv_item = node.get("inventoryItem") or {}
        line_items.append({
            "quantity": node.get("totalQuantity"),
            "shipped": node.get("shippedQuantity", 0),
            "sku": inv_item.get("sku"),
        })

    # Paginar si hay mas
    page_info = li_data.get("pageInfo", {})
    while page_info.get("hasNextPage"):
        cursor = page_info.get("endCursor")
        page_data = _graphql(_TRANSFER_LINE_ITEMS_PAGE, {
            "id": transfer_id, "first": 250, "after": cursor,
        })
        li_data = (page_data.get("inventoryTransfer") or {}).get("lineItems", {})
        for edge in li_data.get("edges", []):
            node = edge["node"]
            inv_item = node.get("inventoryItem") or {}
            line_items.append({
                "quantity": node.get("totalQuantity"),
                "sku": inv_item.get("sku"),
            })
        page_info = li_data.get("pageInfo", {})

    return {
        "id": transfer.get("id"),
        "name": transfer.get("name"),
        "status": transfer.get("status"),
        "date_created": transfer.get("dateCreated"),
        "total_quantity": transfer.get("totalQuantity"),
        "received_quantity": transfer.get("receivedQuantity"),
        "note": transfer.get("note"),
        "reference_name": transfer.get("referenceName"),
        "tags": transfer.get("tags", []),
        "origin": (transfer.get("origin") or {}).get("location", {}).get("name"),
        "destination": (transfer.get("destination") or {}).get("location", {}).get("name"),
        "line_items": line_items,
        "line_items_count": len(line_items),
    }
